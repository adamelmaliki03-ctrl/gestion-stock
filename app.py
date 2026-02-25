import streamlit as st
import pandas as pd
from fpdf import FPDF
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
import os
import hashlib
import numpy as np
from PIL import Image
try:
    import zxingcpp
    QR_DECODE_AVAILABLE = True
except ImportError:
    QR_DECODE_AVAILABLE = False

# ─────────────────────────────────────────────
# CONFIGURATION
# ─────────────────────────────────────────────

st.set_page_config(page_title="GMAO Stock - Campus EMI", layout="wide")

EXCEL_PATH = "stock_campus_emi.xlsx"

# ⚠️ Changer ces identifiants selon vos besoins
USERS = {
    "admin": {
        "password": hashlib.sha256("admin123".encode()).hexdigest(),
        "role": "admin",
        "nom": "Responsable Stock"
    },
    "tech": {
        "password": hashlib.sha256("tech123".encode()).hexdigest(),
        "role": "technicien",
        "nom": "Technicien"
    }
}

# Menus selon le rôle
MENUS_ADMIN = [
    "📦 État du Stock",
    "✏️ Modifier le Stock",
    "📥 Entrée & Facturation",
    "📋 Historique Hebdo",
    "📤 Sortie de Pièce (Scan)"
]
MENUS_TECH = ["📤 Sortie de Pièce (Scan)"]


# ─────────────────────────────────────────────
# FONCTIONS EXCEL
# ─────────────────────────────────────────────

def load_stock_from_excel():
    df = pd.read_excel(EXCEL_PATH, sheet_name="Stock", engine="openpyxl", dtype={"ID_QR": str})
    df["ID_QR"] = df["ID_QR"].astype(str).str.strip().str.rstrip(".0")
    df = df[df["ID_QR"].notna() & (df["ID_QR"] != "TOTAL") & (df["ID_QR"] != "nan")]
    # Remplir les None par des valeurs par défaut
    if "Quantite" not in df.columns:        df["Quantite"] = 0
    if "Prix_Unitaire_DH" not in df.columns: df["Prix_Unitaire_DH"] = 0
    if "Seuil_Alerte" not in df.columns:    df["Seuil_Alerte"] = 0
    df["Quantite"]        = pd.to_numeric(df["Quantite"], errors="coerce").fillna(0).astype(int)
    df["Prix_Unitaire_DH"] = pd.to_numeric(df["Prix_Unitaire_DH"], errors="coerce").fillna(0).astype(float)
    df["Seuil_Alerte"]    = pd.to_numeric(df["Seuil_Alerte"], errors="coerce").fillna(0).astype(int)
    return df[["ID_QR", "Designation", "Quantite", "Prix_Unitaire_DH", "Seuil_Alerte"]].copy()


def save_stock_to_excel(df: pd.DataFrame):
    try:
        wb = load_workbook(EXCEL_PATH)
        ws = wb["Stock"]
        border = Border(left=Side(style="thin"), right=Side(style="thin"),
                        top=Side(style="thin"), bottom=Side(style="thin"))
        alt_fill = PatternFill("solid", start_color="EAF0FB")

        # Effacer les anciennes données
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.value = None

        # Réécrire ligne par ligne (colonnes essentielles uniquement)
        for r_idx, row in enumerate(df.itertuples(index=False), start=2):
            seuil = getattr(row, "Seuil_Alerte", 0) or 0
            values = [str(row.ID_QR), row.Designation, int(row.Quantite),
                      float(row.Prix_Unitaire_DH), f"=C{r_idx}*D{r_idx}", seuil]
            for c_idx, val in enumerate(values, 1):
                cell = ws.cell(r_idx, c_idx, val)
                cell.border = border
                cell.font = Font(name="Arial", size=10)
                cell.alignment = Alignment(horizontal="center" if c_idx != 2 else "left")
                if r_idx % 2 == 0:
                    cell.fill = alt_fill

        # Ligne TOTAL
        total_row = len(df) + 2
        ws.cell(total_row, 1, "TOTAL").font = Font(bold=True, name="Arial")
        ws.cell(total_row, 1).border = border
        total_cell = ws.cell(total_row, 5, f"=SUM(E2:E{total_row-1})")
        total_cell.font = Font(bold=True, name="Arial", color="2E4057")
        total_cell.border = border
        total_cell.alignment = Alignment(horizontal="center")
        for c in [2, 3, 4, 6]:
            ws.cell(total_row, c).border = border

        wb.save(EXCEL_PATH)
    except Exception as e:
        st.error(f"❌ Erreur sauvegarde Excel : {e}")


def ensure_historique_sheet():
    wb = load_workbook(EXCEL_PATH)
    if "Historique_Sorties" not in wb.sheetnames:
        ws2 = wb.create_sheet("Historique_Sorties")
        header_fill = PatternFill("solid", start_color="2E4057")
        header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
        border = Border(left=Side(style="thin"), right=Side(style="thin"),
                        top=Side(style="thin"), bottom=Side(style="thin"))
        for col, (h, w) in enumerate(zip(
            ["Date", "ID_QR", "Designation", "Quantite_Sortie", "Technicien"],
            [22, 12, 35, 18, 25]
        ), 1):
            cell = ws2.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
            ws2.column_dimensions[get_column_letter(col)].width = w
        wb.save(EXCEL_PATH)


def append_sortie_to_excel(date_str, id_qr, designation, qte, technicien):
    wb = load_workbook(EXCEL_PATH)
    ws = wb["Historique_Sorties"]
    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))
    next_row = ws.max_row + 1
    for c_idx, val in enumerate([date_str, id_qr, designation, qte, technicien], 1):
        cell = ws.cell(next_row, c_idx, val)
        cell.border = border
        cell.font = Font(name="Arial", size=10)
        if next_row % 2 == 0:
            cell.fill = PatternFill("solid", start_color="EAF0FB")
    wb.save(EXCEL_PATH)


def load_historique_from_excel():
    return pd.read_excel(EXCEL_PATH, sheet_name="Historique_Sorties", engine="openpyxl")


def to_excel_download(df: pd.DataFrame) -> bytes:
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Historique")
    return output.getvalue()


def generate_pdf(id_trans, fournisseur, items_list, total_general) -> bytes:
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", "B", 16)
    pdf.cell(200, 10, "BON DE RÉCEPTION / FACTURATION STOCK", ln=True, align="C")
    pdf.set_font("Arial", size=12)
    pdf.cell(200, 10, f"Référence : {id_trans} | Date : {datetime.now().strftime('%d/%m/%Y')}", ln=True, align="C")
    pdf.ln(10)
    pdf.cell(100, 10, "Organisme : Campus Universitaire / EMI", ln=True)
    pdf.cell(100, 10, f"Fournisseur : {fournisseur}", ln=True)
    pdf.ln(10)
    pdf.set_fill_color(200, 220, 255)
    pdf.set_font("Arial", "B", 11)
    for col, w in zip(["Désignation", "Qté", "Prix Unitaire", "Total (DH)"], [80, 30, 40, 40]):
        pdf.cell(w, 10, col, border=1, fill=True)
    pdf.ln()
    pdf.set_font("Arial", size=11)
    for item in items_list:
        pdf.cell(80, 10, item["nom"], border=1)
        pdf.cell(30, 10, str(item["qte"]), border=1)
        pdf.cell(40, 10, str(item["prix"]), border=1)
        pdf.cell(40, 10, str(item["total"]), border=1)
        pdf.ln()
    pdf.ln(5)
    pdf.set_font("Arial", "B", 12)
    pdf.cell(150, 10, "TOTAL GÉNÉRAL : ", align="R")
    pdf.cell(40, 10, f"{total_general} DH", border=1, align="C")
    return pdf.output(dest="S").encode("latin-1")


# ─────────────────────────────────────────────
# SESSION STATE
# ─────────────────────────────────────────────

for key, val in [
    ("logged_in", False),
    ("guest_mode", False),
    ("role", None),
    ("username", None),
    ("nom_user", None),
    ("stock_df", None),
]:
    if key not in st.session_state:
        st.session_state[key] = val


# ─────────────────────────────────────────────
# PAGE DE CONNEXION
# ─────────────────────────────────────────────

def page_accueil():
    """Page d'accueil : deux chemins — Technicien (sans mdp) ou Admin (avec mdp)."""
    col_c, col_m, col_c2 = st.columns([1, 2, 1])
    with col_m:
        st.markdown("<br><br>", unsafe_allow_html=True)
        st.markdown("## 🛠️ GMAO Stock - Campus EMI")
        st.markdown("---")
        # Message de confirmation après sortie
        if st.session_state.get("last_sortie_msg"):
            st.success(st.session_state.last_sortie_msg)
            st.session_state.last_sortie_msg = ""

        # ── Bouton Technicien (accès direct) ──
        st.markdown("### 🟢 Technicien")
        st.markdown("Accès direct à la sortie de pièces, sans mot de passe.")
        if st.button("📤 Accéder à la Sortie de Pièce", use_container_width=True):
            st.session_state.guest_mode = True
            st.session_state.role = "technicien"
            st.session_state.nom_user = "Technicien"
            if os.path.exists(EXCEL_PATH):
                st.session_state.stock_df = load_stock_from_excel()
            st.rerun()

        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown("---")

        # ── Connexion Admin ──
        st.markdown("### 🔴 Espace Administrateur")
        username = st.text_input("👤 Identifiant", placeholder="Identifiant admin")
        password = st.text_input("🔑 Mot de passe", type="password", placeholder="Mot de passe")

        if st.button("Se connecter", use_container_width=True, type="primary"):
            if username in USERS and USERS[username]["role"] == "admin":
                hashed = hashlib.sha256(password.encode()).hexdigest()
                if hashed == USERS[username]["password"]:
                    st.session_state.logged_in = True
                    st.session_state.guest_mode = False
                    st.session_state.role = "admin"
                    st.session_state.username = username
                    st.session_state.nom_user = USERS[username]["nom"]
                    st.rerun()
                else:
                    st.error("❌ Mot de passe incorrect.")
            else:
                st.error("❌ Identifiant inconnu ou non autorisé.")

        st.markdown("---")
        st.caption("Projet PFE - EMI Génie Mécanique | Maintenance 4.0")


# ─────────────────────────────────────────────
# APPLICATION PRINCIPALE
# ─────────────────────────────────────────────

def page_app():
    role = st.session_state.role
    nom  = st.session_state.nom_user

    # ── SIDEBAR ──
    st.sidebar.markdown(f"### 👋 Bonjour, **{nom}**")
    badge = "🔴 Admin" if role == "admin" else "🟢 Technicien"
    st.sidebar.markdown(f"Rôle : {badge}")
    st.sidebar.markdown("---")

    # Upload Excel (admin seulement)
    if role == "admin":
        st.sidebar.markdown("### 📂 Charger un fichier Excel")
        uploaded_file = st.sidebar.file_uploader("Déposer votre fichier .xlsx", type=["xlsx"])
        if uploaded_file is not None:
            with open(EXCEL_PATH, "wb") as f:
                f.write(uploaded_file.read())
            df_check = pd.read_excel(EXCEL_PATH, sheet_name=None, engine="openpyxl")
            if "Stock" not in df_check:
                st.sidebar.error("❌ Feuille 'Stock' introuvable.")
            else:
                df_s = df_check["Stock"]
                missing = {"ID_QR", "Designation", "Quantite", "Prix_Unitaire_DH"} - set(df_s.columns)
                if missing:
                    st.sidebar.error(f"❌ Colonnes manquantes : {', '.join(missing)}")
                else:
                    ensure_historique_sheet()
                    st.session_state.stock_df = load_stock_from_excel()
                    st.sidebar.success(f"✅ Fichier chargé : {uploaded_file.name}")

        if st.sidebar.button("🔄 Recharger depuis Excel"):
            if os.path.exists(EXCEL_PATH):
                st.session_state.stock_df = load_stock_from_excel()
                st.sidebar.success("Rechargé !")
        st.sidebar.markdown("---")

    # Menu selon rôle
    menus = MENUS_ADMIN if role == "admin" else MENUS_TECH
    menu  = st.sidebar.radio("Navigation", menus)

    st.sidebar.markdown("---")
    # Bouton selon le mode
    if st.session_state.guest_mode:
        if st.sidebar.button("🏠 Retour à l'accueil"):
            st.session_state.guest_mode = False
            st.session_state.role = None
            st.session_state.nom_user = None
            st.rerun()
    else:
        if st.sidebar.button("🚪 Se déconnecter"):
            for key in ["logged_in", "guest_mode", "role", "username", "nom_user"]:
                st.session_state[key] = False if key in ["logged_in", "guest_mode"] else None
            st.rerun()

    st.sidebar.markdown("---")
    st.sidebar.info("Projet PFE - EMI Génie Mécanique\nMaintenance 4.0")

    # ── TITRE ──
    st.title("🛠️ Gestion de Stock & Maintenance - Campus EMI")

    # ── GARDE : chargement initial du stock (une seule fois) ──
    if st.session_state.stock_df is None:
        if os.path.exists(EXCEL_PATH):
            st.session_state.stock_df = load_stock_from_excel()
        elif role == "admin":
            st.info("👈 **Chargez votre fichier Excel** via la barre latérale pour commencer.")
            st.markdown("""
**Colonnes requises dans la feuille `Stock` :**

| ID_QR | Designation | Quantite | Prix_Unitaire_DH | Seuil_Alerte *(optionnel)* |
|---|---|---|---|---|
| 222 | Nom de la pièce | 10 | 500 | 3 |
            """)
            st.stop()
        else:
            st.warning("⚠️ Aucun stock disponible. Contactez l'administrateur.")
            st.stop()

    # ════════════════════════════════════════
    # ONGLET : ÉTAT DU STOCK  (admin)
    # ════════════════════════════════════════
    if menu == "📦 État du Stock":
        st.subheader("Inventaire des pièces de rechange")
        col1, col2 = st.columns([3, 1])
        with col1:
            df_display = st.session_state.stock_df.copy()
            df_display["Valeur_Totale_DH"] = df_display["Quantite"] * df_display["Prix_Unitaire_DH"]

            def highlight_low(row):
                seuil = row.get("Seuil_Alerte", 0) or 0
                color = "background-color: #FFD6D6" if row["Quantite"] <= seuil else ""
                return [color] * len(row)

            st.dataframe(df_display.style.apply(highlight_low, axis=1), use_container_width=True)
            st.caption("🔴 Fond rouge = quantité ≤ seuil d'alerte")
        with col2:
            st.metric("Nb références", len(df_display))
            total_val = int((df_display["Quantite"] * df_display["Prix_Unitaire_DH"]).sum())
            st.metric("Valeur totale stock", f"{total_val:,} DH")
        st.divider()
        with open(EXCEL_PATH, "rb") as f:
            st.download_button(
                label="📥 Télécharger le fichier Excel",
                data=f.read(),
                file_name="stock_campus_emi.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    # ════════════════════════════════════════
    # ONGLET : MODIFIER LE STOCK  (admin)
    # ════════════════════════════════════════
    elif menu == "✏️ Modifier le Stock":
        st.subheader("Modifier / Corriger le stock")

        tab1, tab2, tab3 = st.tabs(["✏️ Modifier une pièce", "➕ Ajouter une pièce", "🗑️ Supprimer une pièce"])

        # ── Modifier ──
        with tab1:
            df = st.session_state.stock_df
            id_mod = st.selectbox("Sélectionner la pièce à modifier", df["ID_QR"], key="mod_id")
            idx = df[df["ID_QR"] == id_mod].index[0]
            row = df.loc[idx]

            with st.form("form_modifier"):
                new_designation = st.text_input("Désignation", value=str(row["Designation"]))
                new_qte         = st.number_input("Quantité", min_value=0, value=int(row["Quantite"]))
                new_prix        = st.number_input("Prix Unitaire (DH)", min_value=0.0, value=float(row["Prix_Unitaire_DH"]), step=0.5)
                new_seuil       = st.number_input("Seuil d'alerte", min_value=0, value=int(row["Seuil_Alerte"] or 0))
                submit_mod      = st.form_submit_button("💾 Enregistrer les modifications", type="primary")

            if submit_mod:
                st.session_state.stock_df.at[idx, "Designation"]     = new_designation
                st.session_state.stock_df.at[idx, "Quantite"]        = new_qte
                st.session_state.stock_df.at[idx, "Prix_Unitaire_DH"]= new_prix
                st.session_state.stock_df.at[idx, "Seuil_Alerte"]    = new_seuil
                save_stock_to_excel(st.session_state.stock_df)
                st.success(f"✅ Pièce **{id_mod}** mise à jour et sauvegardée dans Excel.")

        # ── Ajouter ──
        with tab2:
            with st.form("form_ajouter"):
                new_id    = st.text_input("ID QR (ex: PMP-06)", placeholder="XXX-00")
                new_des   = st.text_input("Désignation")
                new_q     = st.number_input("Quantité initiale", min_value=0, value=1)
                new_p     = st.number_input("Prix Unitaire (DH)", min_value=0.0, value=0.0, step=0.5)
                new_s     = st.number_input("Seuil d'alerte", min_value=0, value=0)
                submit_aj = st.form_submit_button("➕ Ajouter la pièce", type="primary")

            if submit_aj:
                df = st.session_state.stock_df
                if new_id.strip() == "":
                    st.error("❌ L'ID QR ne peut pas être vide.")
                elif new_id in df["ID_QR"].values:
                    st.error(f"❌ L'ID **{new_id}** existe déjà.")
                else:
                    nouvelle_ligne = pd.DataFrame([{
                        "ID_QR": new_id, "Designation": new_des,
                        "Quantite": new_q, "Prix_Unitaire_DH": new_p, "Seuil_Alerte": new_s
                    }])
                    st.session_state.stock_df = pd.concat(
                        [st.session_state.stock_df, nouvelle_ligne], ignore_index=True
                    )
                    save_stock_to_excel(st.session_state.stock_df)
                    st.success(f"✅ Pièce **{new_id}** ajoutée avec succès.")

        # ── Supprimer ──
        with tab3:
            df = st.session_state.stock_df
            id_del = st.selectbox("Sélectionner la pièce à supprimer", df["ID_QR"], key="del_id")
            st.warning(f"⚠️ Vous allez supprimer définitivement **{id_del}** du stock.")
            if st.button("🗑️ Confirmer la suppression", type="primary"):
                st.session_state.stock_df = df[df["ID_QR"] != id_del].reset_index(drop=True)
                save_stock_to_excel(st.session_state.stock_df)
                st.success(f"✅ Pièce **{id_del}** supprimée et Excel mis à jour.")
                st.rerun()

    # ════════════════════════════════════════
    # ONGLET : ENTRÉE & FACTURATION  (admin)
    # ════════════════════════════════════════
    elif menu == "📥 Entrée & Facturation":
        st.subheader("Réception de commande & Génération de facture")
        with st.form("form_entree"):
            fournisseur = st.text_input("Nom du Fournisseur")
            id_piece    = st.selectbox("Sélectionner la pièce reçue", st.session_state.stock_df["ID_QR"])
            qte_entree  = st.number_input("Quantité reçue", min_value=1, value=1)
            valider     = st.form_submit_button("Enregistrer l'Entrée & Préparer Facture", type="primary")

        if valider:
            df  = st.session_state.stock_df
            idx = df[df["ID_QR"] == id_piece].index[0]
            st.session_state.stock_df.at[idx, "Quantite"] += qte_entree
            nom_p  = df.at[idx, "Designation"]
            prix_p = df.at[idx, "Prix_Unitaire_DH"]
            save_stock_to_excel(st.session_state.stock_df)
            st.success(f"✅ Stock mis à jour pour **{nom_p}**. Excel sauvegardé.")
            items_pdf = [{"nom": nom_p, "qte": qte_entree, "prix": prix_p, "total": qte_entree * prix_p}]
            pdf_bytes = generate_pdf(
                f"FAC-{datetime.now().strftime('%H%M%S')}", fournisseur, items_pdf, qte_entree * prix_p
            )
            st.download_button(
                label="📄 Télécharger la Feuille de Facturation (PDF)",
                data=pdf_bytes,
                file_name=f"facture_{id_piece}.pdf",
                mime="application/pdf"
            )

    # ════════════════════════════════════════
    # ONGLET : HISTORIQUE HEBDO  (admin)
    # ════════════════════════════════════════
    elif menu == "📋 Historique Hebdo":
        st.subheader("Pièces sorties pendant la semaine")
        df_hist = load_historique_from_excel()
        if df_hist.empty or df_hist.dropna(how="all").empty:
            st.info("Aucune sortie enregistrée pour le moment.")
        else:
            df_hist["Date_dt"] = pd.to_datetime(df_hist["Date"], errors="coerce")
            df_hebdo = df_hist[
                df_hist["Date_dt"] > datetime.now() - timedelta(days=7)
            ].drop(columns=["Date_dt"])
            if df_hebdo.empty:
                st.info("Aucune sortie cette semaine.")
            else:
                st.dataframe(df_hebdo, use_container_width=True)
                st.metric("Total sorties cette semaine", len(df_hebdo))
                st.download_button(
                    label="📊 Exporter vers Excel",
                    data=to_excel_download(df_hebdo),
                    file_name=f"rapport_hebdo_{datetime.now().strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

    # ════════════════════════════════════════
    # ONGLET : SORTIE DE PIÈCE  (tous)
    # ════════════════════════════════════════
    elif menu == "📤 Sortie de Pièce (Scan)":
        st.subheader("Sortie de matériel par Scan QR")

        # ── Initialisation de l'ID scanné en session ──
        if "scanned_id" not in st.session_state:
            st.session_state.scanned_id = ""

        # ── Caméra + décodage automatique du QR ──
        img_file = st.camera_input("📷 Scanner le QR Code sur la pièce")

        if img_file is not None:
            if QR_DECODE_AVAILABLE:
                img = Image.open(img_file).convert("RGB")
                img_np = np.array(img)
                results = zxingcpp.read_barcodes(img_np)
                if results:
                    decoded = results[0].text.strip().replace(" ", "")
                    st.session_state.scanned_id = decoded
                    st.success(f"✅ QR Code détecté : **{decoded}**")
                else:
                    st.warning("⚠️ QR Code non lisible. Rapprochez la caméra ou saisissez l'ID manuellement.")
            else:
                st.warning("⚠️ Décodage non disponible. Saisissez l'ID manuellement.")

        # ── Champ ID : pré-rempli si QR scanné ──
        id_scan = st.text_input(
            "🔢 ID de la pièce",
            value=st.session_state.scanned_id,
            placeholder="Ex: PMP-01"
        )
        # Nettoyage et synchronisation
        id_scan = id_scan.strip().replace(" ", "")
        st.session_state.scanned_id = id_scan

        # Aperçu de la pièce si l'ID est reconnu
        df = st.session_state.stock_df
        # Normalisation des IDs du stock pour comparaison robuste
        df_ids = df["ID_QR"].astype(str).str.strip()
        if id_scan and id_scan in df_ids.values:
            idx = df[df_ids == id_scan].index[0]
            nom = df.at[idx, "Designation"]
            st.info(f"🔩 **{nom}**")
        elif id_scan:
            st.error(f"❌ Pièce '{id_scan}' non trouvée dans la base de données.")

        qte_sortie = st.number_input("Quantité à retirer", min_value=1, value=1)

        # Nom du technicien : pré-rempli si connecté
        default_name = st.session_state.nom_user or ""
        user_name = st.text_input("Nom du technicien", value=default_name)

        if "last_sortie_msg" in st.session_state and st.session_state.last_sortie_msg:
            st.success(st.session_state.last_sortie_msg)
            st.session_state.last_sortie_msg = ""

        if st.button("✅ Valider la Sortie", type="primary"):
            # Relire le stock frais depuis la session au moment du clic
            df_live = st.session_state.stock_df.copy()
            df_live_ids = df_live["ID_QR"].astype(str).str.strip()
            id_val = st.session_state.scanned_id.strip()

            if not id_val:
                st.warning("⚠️ Veuillez scanner ou saisir un ID.")
            elif id_val in df_live_ids.values:
                idx = df_live[df_live_ids == id_val].index[0]
                stock_actuel = int(df_live.at[idx, "Quantite"])
                designation  = df_live.at[idx, "Designation"]
                if stock_actuel >= qte_sortie:
                    # Mise à jour directe sur st.session_state.stock_df
                    st.session_state.stock_df.at[idx, "Quantite"] = stock_actuel - qte_sortie
                    # Sauvegarde Excel
                    save_stock_to_excel(st.session_state.stock_df)
                    # Historique
                    append_sortie_to_excel(
                        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        id_val, designation, qte_sortie, user_name
                    )
                    st.session_state.scanned_id = ""
                    st.session_state.guest_mode = False
                    st.session_state.role = None
                    st.session_state.nom_user = None
                    st.session_state["last_sortie_msg"] = f"✅ Sortie validée : {qte_sortie} × {designation} retiré(s) par {user_name}."
                    st.rerun()
                else:
                    st.error(f"❌ Stock insuffisant ! Stock actuel : {stock_actuel}")
            else:
                st.warning(f"⚠️ Pièce '{id_val}' non trouvée dans la base de données.")


# ─────────────────────────────────────────────
# ROUTAGE PRINCIPAL
# ─────────────────────────────────────────────

if not st.session_state.logged_in and not st.session_state.guest_mode:
    page_accueil()
else:
    page_app()
